"""Leitura das planilhas oficiais da Matriz CONIF (.xlsx) — substitui a automação COM do Excel
via PowerShell usada antes (ver reference_planilhas_oficiais_conif.md), agora que Python está
instalado no ambiente.

Por que este é o jeito mais eficiente/recomendado para este caso (arquivo grande, cheio de
fórmulas, só leitura, ver Metodologia_Matriz_Orcamentaria_CONIF.md):

  - `read_only=True`: openpyxl entra em modo streaming (lê a XML da planilha sequencialmente em
    vez de carregar a árvore de objetos inteira em memória) — é a própria recomendação oficial do
    openpyxl para "só preciso ler dados". Muito mais rápido e leve que o modo padrão, e mais
    rápido que abrir o Excel de verdade via COM (que precisa da aplicação inteira carregada).
  - `data_only=True`: retorna o VALOR já calculado da fórmula (o cache que o Excel salva junto do
    arquivo), não a string da fórmula (ex.: em vez de "=I25*(1+I10)", devolve 2951996879.71).

  ATENÇÃO — a única pegadinha real: o cache de `data_only=True` só existe se o arquivo foi aberto
  e salvo pelo Excel de verdade depois da última edição das fórmulas. openpyxl (e este script)
  NUNCA calculam fórmula nenhuma sozinhos — só leem o que já está salvo. Se uma célula que
  deveria ter fórmula vier `None`, abra o arquivo no Excel, deixe recalcular (F9) e salve antes de
  rodar de novo.

Uso: `python scripts/ler_planilha_conif.py` roda o exemplo no fim do arquivo contra a planilha
2026 conhecida. Para outro ano/arquivo, chamar `abrir_planilha()`/`ler_celula()`/
`ler_aba_como_dataframe()` a partir de outro script ou do REPL, apontando para o caminho certo.
"""

from pathlib import Path

import openpyxl
import pandas as pd

CAMINHO_PLANILHA_2026 = Path(
    r"C:\Users\USER\OneDrive\Documentos\IFSul\_Matriz orçamentária - CONIF\2026\Mauro - NH\5ª Fase - MATRIZ CONIF 2026 (1).xlsx"
)


def abrir_planilha(caminho: Path) -> openpyxl.Workbook:
    """Abre o workbook em modo streaming + só-valores-calculados (ver módulo acima)."""
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")
    return openpyxl.load_workbook(caminho, data_only=True, read_only=True)


def ler_celula(wb: openpyxl.Workbook, aba: str, endereco: str):
    """Lê uma única célula por endereço estilo Excel (ex.: 'I10') — para os parâmetros soltos de
    DADOS BASE (IPCA, valor da matriz do ano anterior etc., ver Metodologia_Matriz_Orcamentaria_CONIF.md § 2)."""
    return wb[aba][endereco].value


def ler_aba_como_dataframe(
    wb: openpyxl.Workbook,
    aba: str,
    linha_cabecalho: int,
    linha_inicio_dados: int,
) -> pd.DataFrame:
    """Extrai uma aba tabular inteira para DataFrame, numa única passada sequencial pela planilha
    (compatível com o modo streaming de `read_only=True` — nunca faz `seek` para trás).

    `linha_cabecalho` e `linha_inicio_dados` são 1-indexados, como no próprio Excel, e não
    precisam ser adjacentes: em COMPLETO PROPOSTA, por exemplo, o cabeçalho fica na linha 8 mas os
    dados só começam na 14 (linhas 9-13 são outra coisa — título/legenda) — ver
    reference_planilhas_oficiais_conif.md.
    """
    planilha = wb[aba]
    cabecalho: tuple | None = None
    dados: list[tuple] = []

    for i, linha in enumerate(planilha.iter_rows(values_only=True), start=1):
        if i == linha_cabecalho:
            cabecalho = linha
        elif i >= linha_inicio_dados:
            dados.append(linha)

    if cabecalho is None:
        raise ValueError(f"Linha de cabeçalho {linha_cabecalho} não encontrada na aba '{aba}'.")

    return pd.DataFrame(dados, columns=cabecalho)


if __name__ == "__main__":
    wb = abrir_planilha(CAMINHO_PLANILHA_2026)

    print("Abas disponíveis:", wb.sheetnames)

    # Exemplo 1: parâmetro solto de DADOS BASE (Metodologia_Matriz_Orcamentaria_CONIF.md § 2).
    print("DADOS BASE!I10 (IPCA):", ler_celula(wb, "DADOS BASE", "I10"))
    print("DADOS BASE!I25 (Valor da Matriz do ano anterior):", ler_celula(wb, "DADOS BASE", "I25"))

    # Exemplo 2: aba tabular inteira como DataFrame (mesmo layout de COMPLETO PROPOSTA já
    # documentado: cabeçalho na linha 8, dados a partir da 14).
    df = ler_aba_como_dataframe(wb, "COMPLETO PROPOSTA", linha_cabecalho=8, linha_inicio_dados=14)
    print(f"COMPLETO PROPOSTA: {len(df)} linha(s), colunas: {list(df.columns)[:8]}...")
    print(df.head(3))
